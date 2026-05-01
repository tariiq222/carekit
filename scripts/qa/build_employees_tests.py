"""
Build E2E Employees sheet — 120 test cases, organized by category.
Status values: Pass | Fail | Blocked | Not Started
"""
import openpyxl
from openpyxl.styles import PatternFill, Alignment, Border, Side
from copy import copy
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

WB_PATH = "e2e_test_clients.xlsx"
SHEET = "E2E Employees"
MODULE = "Employees (الممارسون)"
TODAY = "2026-04-15"
ENV = "Local dev"
EXEC = "Claude"

HEADERS = [
    "Test ID", "Module", "Test Type", "Test Scenario", "Test Case Title",
    "Preconditions", "Test Steps", "Test Data", "Expected Result",
    "Actual Result", "Priority", "Status", "Environment", "Executed By",
    "Execution Date", "Notes / Defect ID",
]

# Each row: (id, type, scenario, title, pre, steps, data, expected, actual, priority, status, notes)
# Module/Env/Exec/Date auto-filled. Actual/Status/Notes empty => Not Started.
T = []

def add(tid, ttype, scenario, title, pre, steps, data, expected,
        actual="", priority="P2-Medium", status="Not Started", notes=""):
    T.append((tid, ttype, scenario, title, pre, steps, data, expected,
              actual, priority, status, notes))

# ═══════════════════════════════════════════════════════════════════════════
# CATEGORY 1: LIST + PAGINATION + SORTING  (EM-001 → EM-015)
# ═══════════════════════════════════════════════════════════════════════════
add("EM-001", "Positive", "عرض القائمة", "تحميل قائمة الممارسين الأولية",
    "مستخدم مسجّل + ممارسون في tenant",
    "افتح /employees",
    "page=1&limit=20",
    "GET /employees → 200 + GET /stats → 200. عرض الجدول + StatsGrid.",
    "نجح: endpoints منفصلة تعمل.",
    "P1-High", "Pass",
    "مُختبَر اليوم.")

add("EM-002", "Positive", "Pagination", "الانتقال لصفحة 2",
    ">20 ممارس في DB",
    "1. لاحظ pagination bar\n2. اضغط (التالي)",
    "page=2",
    "GET ?page=2&limit=20 → 200. عرض السجلات 21-40.",
    "", "P1-High", "Not Started",
    "لا يمكن الاختبار: DB فيها <20 ممارس.")

add("EM-003", "Positive", "Pagination", "تغيير limit",
    "قائمة محمّلة",
    "1. غيّر limit لـ 50\n2. لاحظ طلب الشبكة",
    "limit=50",
    "GET ?limit=50. عرض حتى 50 سجل/صفحة.",
    "", "P2-Medium", "Not Started",
    "UI للـ limit-picker غير ظاهر في الصفحة الحالية.")

add("EM-004", "Positive", "Sorting", "فرز باسم الممارس (تصاعدي)",
    "≥2 ممارسين",
    "اضغط header (الممارس)",
    "sortBy=name&sortOrder=asc",
    "البيانات مرتبة أبجدياً",
    "", "P2-Medium", "Not Started",
    "يحتاج بيانات متعددة.")

add("EM-005", "Positive", "Sorting", "فرز بالخبرة (تنازلي)",
    "ممارسون بخبرات مختلفة",
    "اضغط header (الخبرة) مرتين",
    "sortBy=experience&sortOrder=desc",
    "الأعلى خبرة أولاً",
    "", "P2-Medium", "Not Started", "")

add("EM-006", "Positive", "Sorting", "فرز بالتقييم",
    "ممارسون بتقييمات مختلفة",
    "اضغط header (التقييم)",
    "sortBy=averageRating",
    "الأعلى تقييماً أولاً",
    "", "P3-Low", "Not Started", "")

add("EM-007", "Positive", "Stats", "StatsGrid تعكس الكل بلا فلتر",
    "قائمة مملوءة",
    "افتح الصفحة",
    "—",
    "الإجمالي = عدد كل الممارسين، وليس النتيجة المفلترة",
    "نجح بعد إضافة endpoint stats منفصل.",
    "P1-High", "Pass",
    "Fixed: employee-stats.handler.ts")

add("EM-008", "Positive", "Stats", "StatsGrid لا تتأثر بالبحث",
    "قائمة مملوءة",
    "1. ابحث عن نص غير موجود\n2. لاحظ الكروت",
    "search=zzz",
    "Stats ثابتة",
    "نجح — Stats مستقلة.",
    "P1-High", "Pass", "")

add("EM-009", "Positive", "Empty state", "لا يوجد ممارسون",
    "DB فارغة من الممارسين",
    "افتح /employees",
    "—",
    "عرض (لا يوجد ممارسون حتى الآن)",
    "نجح — empty state واضح.",
    "P2-Medium", "Pass", "")

add("EM-010", "Negative", "Error", "backend down → رسالة خطأ",
    "Backend متوقف",
    "أوقف الباك واحمّل الصفحة",
    "—",
    "ErrorBanner يظهر، لا crash",
    "", "P2-Medium", "Not Started",
    "لم أختبر — الباك شغّال.")

add("EM-011", "Positive", "Columns", "عرض كل الأعمدة المطلوبة",
    "صف واحد على الأقل",
    "افحص header الجدول",
    "—",
    "Columns: الممارس, البريد, الخبرة, التقييم, الحالة, الإجراءات",
    "نجح.", "P2-Medium", "Pass", "")

add("EM-012", "Positive", "Cell rendering", "عرض الخبرة بلاحقة صحيحة",
    "ممارس بخبرة 15",
    "لاحظ خانة الخبرة",
    "experience=15",
    "(15 سنة) في AR — (15 yrs) في EN",
    "نجح.", "P3-Low", "Pass", "")

add("EM-013", "Positive", "Cell rendering", "عرض التقييم عند null",
    "ممارس بلا تقييمات",
    "لاحظ خانة التقييم",
    "averageRating=null",
    "يعرض — (em dash)",
    "نجح.", "P3-Low", "Pass", "")

add("EM-014", "Positive", "Badge", "Badge نشط أخضر",
    "ممارس isActive=true",
    "لاحظ خانة الحالة",
    "—",
    "Badge بخلفية success/10 ونص success",
    "نجح.", "P3-Low", "Pass", "")

add("EM-015", "Positive", "Badge", "Badge غير نشط رمادي",
    "ممارس isActive=false",
    "لاحظ خانة الحالة",
    "—",
    "Badge رمادي مع نص معطّل",
    "", "P3-Low", "Not Started",
    "DB فيها ممارسين نشطين فقط.")

# ═══════════════════════════════════════════════════════════════════════════
# CATEGORY 2: SEARCH  (EM-020 → EM-030)
# ═══════════════════════════════════════════════════════════════════════════
add("EM-020", "Positive", "Search", "بحث باسم كامل",
    "ممارس اسمه معروف",
    "اكتب (tariq) في البحث",
    "search=tariq",
    "الجدول يعرض المطابق فقط",
    "نجح.", "P1-High", "Pass", "")

add("EM-021", "Positive", "Search Debounce", "بحث سريع = طلب واحد",
    "قائمة مملوءة",
    "اكتب 5 أحرف سريعة",
    "tariq",
    "طلب واحد بعد 300ms",
    "نجح — قبل الإصلاح 5 طلبات، الآن 1.",
    "P2-Medium", "Pass",
    "Fixed: use-employees.ts debounce")

add("EM-022", "Positive", "Search", "بحث بجزء من الاسم",
    "ممارس اسمه أحمد",
    "اكتب (أح)",
    "search=أح",
    "يطابق (أحمد) و أي اسم يحتوي أح",
    "", "P2-Medium", "Not Started", "")

add("EM-023", "Positive", "Search", "بحث بالإيميل",
    "ممارس بإيميل معروف",
    "اكتب جزء من الإيميل",
    "search=test.com",
    "يطابق الإيميل",
    "", "P2-Medium", "Not Started",
    "تحقّق backend WHERE يبحث في email.")

add("EM-024", "Positive", "Search", "بحث بالجوال",
    "ممارس برقم جوال",
    "اكتب جزء من الرقم",
    "search=0501",
    "يطابق",
    "", "P2-Medium", "Not Started", "")

add("EM-025", "Negative", "Search", "بحث بنص غير موجود",
    "قائمة مملوءة",
    "اكتب نص عشوائي",
    "search=zzznotfoundxyz",
    "Empty state + Stats ثابتة",
    "نجح.", "P2-Medium", "Pass", "")

add("EM-026", "Edge case", "Search", "بحث عربي + إنجليزي",
    "ممارسون باسمين",
    "اكتب (أحمد) ثم (Ahmed)",
    "—",
    "كلاهما يعمل",
    "", "P3-Low", "Not Started", "")

add("EM-027", "Security", "Search", "حقن SQL في البحث",
    "—",
    "اكتب (' OR 1=1 --)",
    "search=' OR 1=1 --",
    "لا تنفيذ SQL، Prisma يتعامل معه كـ string",
    "", "P1-High", "Not Started",
    "نظري: Prisma parameterized queries آمنة. يحتاج pen-test.")

add("EM-028", "Edge case", "Search", "بحث بأحرف خاصة",
    "—",
    "اكتب (%$&*)",
    "—",
    "لا crash، يرجع 200 + قائمة فارغة",
    "", "P3-Low", "Not Started", "")

add("EM-029", "Edge case", "Search", "بحث بمسافات فقط",
    "—",
    "اكتب 5 مسافات",
    "search=%20%20%20%20%20",
    "trim على الباك، يعامل كفارغ",
    "", "P3-Low", "Not Started", "")

add("EM-030", "Edge case", "Search", "بحث طويل جداً (>200 char)",
    "—",
    "الصق نص 500 حرف",
    "—",
    "الباك يقبل أو يرد 400 برسالة واضحة",
    "", "P3-Low", "Not Started", "")

# ═══════════════════════════════════════════════════════════════════════════
# CATEGORY 3: FILTERS  (EM-031 → EM-040)
# ═══════════════════════════════════════════════════════════════════════════
add("EM-031", "Positive", "Filter", "فلتر نشط",
    "ممارسون بحالات مختلطة",
    "اختر (نشط) من combobox",
    "isActive=true",
    "عرض النشطين فقط",
    "نجح.", "P1-High", "Pass",
    "Fixed: controller parses rawIsActive manually.")

add("EM-032", "Positive", "Filter", "فلتر غير نشط",
    "ممارس معطّل موجود",
    "اختر (غير متاح)",
    "isActive=false",
    "عرض المعطّلين فقط",
    "نجح.", "P1-High", "Pass", "")

add("EM-033", "Positive", "Filter", "Reset فلتر",
    "فلتر مُطبَّق",
    "اضغط (إعادة تعيين)",
    "—",
    "كل الفلاتر تختفي + القائمة كاملة",
    "نجح.", "P2-Medium", "Pass", "")

add("EM-034", "Positive", "Filter", "جميع الحالات",
    "فلتر نشط مُطبَّق",
    "اختر (جميع الحالات)",
    "isActive=undefined",
    "param isActive لا يُرسل",
    "", "P2-Medium", "Not Started", "")

add("EM-035", "Positive", "Filter", "Combine search + status",
    "قائمة مختلطة",
    "اكتب بحث + اختر حالة",
    "search=ahmed&isActive=true",
    "كلا الفلترين يُطبَّقان",
    "", "P2-Medium", "Not Started", "")

add("EM-036", "Positive", "Filter", "Filter + pagination reset",
    "في صفحة 2",
    "طبّق فلتر",
    "—",
    "العودة لصفحة 1 تلقائياً",
    "", "P2-Medium", "Not Started", "")

add("EM-037", "Positive", "Filter", "Filter persists على reload",
    "فلتر مُطبَّق",
    "اضغط F5",
    "—",
    "الفلتر يبقى (via URL أو state)",
    "", "P3-Low", "Not Started",
    "حالياً الفلتر لا يبقى بعد reload.")

add("EM-038", "Positive", "Filter counter", "Reset zr يظهر فقط عند فلاتر",
    "لا فلاتر",
    "لاحظ زر Reset",
    "—",
    "مخفي",
    "نجح.", "P3-Low", "Pass", "")

add("EM-039", "Positive", "Filter counter", "عدد النتائج في FilterBar",
    "فلتر يعطي 3 نتائج",
    "لاحظ شريط الفلاتر",
    "—",
    "يعرض (3 الإجمالي)",
    "نجح.", "P3-Low", "Pass", "")

add("EM-040", "Edge case", "Filter", "ترشيح مع 0 نتائج",
    "—",
    "طبّق فلتر لا يطابق شيء",
    "—",
    "Empty state + Stats ثابتة",
    "نجح.", "P2-Medium", "Pass", "")

# ═══════════════════════════════════════════════════════════════════════════
# CATEGORY 4: CREATE — Basic Info  (EM-041 → EM-055)
# ═══════════════════════════════════════════════════════════════════════════
add("EM-041", "Positive", "Create", "إنشاء بالحد الأدنى",
    "صلاحية create",
    "املأ nameEn + nameAr + email + specialty + experience → احفظ",
    "Dr. Test/د. اختبار/test@test.com/Testing/5",
    "POST /onboarding → 201",
    "نجح.", "P1-High", "Pass", "")

add("EM-042", "Positive", "Create", "إنشاء بكل الحقول",
    "—",
    "املأ كل الحقول (title, bio, education, avatarUrl...)",
    "كل الحقول",
    "201 + كل الحقول محفوظة",
    "نجح.", "P1-High", "Pass", "")

add("EM-043", "Negative", "Validation", "nameEn فارغ",
    "—",
    "اترك الاسم الإنجليزي فارغ",
    "nameEn=''",
    "رسالة خطأ + لا طلب POST",
    "", "P1-High", "Not Started",
    "تحقّق Zod schema + RHF.")

add("EM-044", "Negative", "Validation", "nameAr فارغ",
    "—",
    "اترك الاسم العربي فارغ",
    "—",
    "رسالة خطأ",
    "", "P1-High", "Not Started", "")

add("EM-045", "Negative", "Validation", "email فارغ",
    "—",
    "اترك الإيميل فارغ",
    "—",
    "رسالة خطأ",
    "", "P1-High", "Not Started", "")

add("EM-046", "Negative", "Validation", "email بصيغة خطأ",
    "—",
    "أدخل (not-email)",
    "email=not-email",
    "رسالة خطأ في الفورم",
    "", "P1-High", "Not Started", "")

add("EM-047", "Negative", "Validation", "email مكرّر",
    "ممارس بإيميل سابق",
    "أنشئ ممارس بنفس الإيميل",
    "—",
    "400 أو 409",
    "", "P1-High", "Not Started",
    "تحقّق constraint @@unique([tenantId, email]).")

add("EM-048", "Negative", "Validation", "specialty فارغ",
    "—",
    "اترك التخصص EN فارغ",
    "—",
    "رسالة خطأ (حقل required)",
    "", "P1-High", "Not Started", "")

add("EM-049", "Negative", "Validation", "experience سالب",
    "—",
    "أدخل -5",
    "experience=-5",
    "رسالة خطأ (min=0)",
    "", "P2-Medium", "Not Started", "")

add("EM-050", "Negative", "Validation", "experience عشري",
    "—",
    "أدخل 5.5",
    "experience=5.5",
    "يُقبل كـ int (5) أو رفض",
    "", "P3-Low", "Not Started", "")

add("EM-051", "Negative", "Validation", "experience غير عددي",
    "—",
    "أدخل (abc)",
    "—",
    "HTML input type=number يمنع",
    "", "P3-Low", "Not Started", "")

add("EM-052", "Edge case", "Validation", "nameAr طويل >255",
    "—",
    "الصق 500 حرف",
    "—",
    "رفض أو truncate",
    "", "P3-Low", "Not Started",
    "تحقّق DB column length.")

add("EM-053", "Positive", "i18n in DB", "إنشاء بأسماء Unicode",
    "—",
    "أدخل nameAr بحروف خاصة + emoji",
    "nameAr=د. 🩺 أحمد",
    "يُحفظ كما هو",
    "", "P3-Low", "Not Started", "")

add("EM-054", "Positive", "Create UX", "Cancel يعيد للقائمة",
    "فورم مملوء",
    "اضغط (إلغاء)",
    "—",
    "لا حفظ + redirect",
    "نجح.", "P3-Low", "Pass", "")

add("EM-055", "Positive", "Create UX", "isActive default true",
    "فورم جديد",
    "لاحظ switch",
    "—",
    "switch مفعّل افتراضياً",
    "نجح.", "P2-Medium", "Pass", "")

# ═══════════════════════════════════════════════════════════════════════════
# CATEGORY 5: CREATE — Schedule  (EM-056 → EM-070)
# ═══════════════════════════════════════════════════════════════════════════
add("EM-056", "Positive", "Schedule", "أيام افتراضية",
    "tab 2 مفتوح",
    "لاحظ الـ switches",
    "—",
    "الأحد-الخميس مفعّلة، الجمعة/السبت معطّلة",
    "نجح.", "P2-Medium", "Pass", "")

add("EM-057", "Positive", "Schedule", "تعطيل يوم",
    "—",
    "عطّل الأحد",
    "—",
    "dayOfWeek=0 لا يُرسل في PATCH",
    "نجح.", "P1-High", "Pass", "")

add("EM-058", "Positive", "Schedule", "تفعيل السبت",
    "السبت معطّل",
    "فعّل السبت",
    "—",
    "dayOfWeek=6 يُرسل",
    "نجح.", "P1-High", "Pass", "")

add("EM-059", "Positive", "Schedule", "كل الأيام مفعّلة",
    "—",
    "فعّل السبت + الجمعة",
    "—",
    "7 windows",
    "", "P2-Medium", "Not Started", "")

add("EM-060", "Negative", "Schedule", "كل الأيام معطّلة",
    "—",
    "عطّل كل يوم",
    "activeSlots=[]",
    "PATCH availability لا يُرسل (حسب الكود الحالي)",
    "", "P2-Medium", "Not Started",
    "if (activeSlots.length > 0) check يمنع الطلب.")

add("EM-061", "Positive", "Schedule", "تغيير ساعات يوم",
    "الأحد مفعّل",
    "غيّر 09:00 → 10:00",
    "startTime=10:00",
    "PATCH يرسل القيمة الجديدة",
    "", "P1-High", "Not Started",
    "فشل اختباري أمس: spinbutton صعب التفاعل.")

add("EM-062", "Negative", "Schedule", "endTime قبل startTime",
    "الأحد مفعّل",
    "ضع 17:00 → 09:00",
    "start=17:00, end=09:00",
    "رفض أو رسالة خطأ",
    "", "P1-High", "Not Started",
    "تحقّق validation.")

add("EM-063", "Negative", "Schedule", "endTime = startTime",
    "—",
    "ضع 09:00 → 09:00",
    "—",
    "رفض",
    "", "P2-Medium", "Not Started", "")

add("EM-064", "Edge case", "Schedule", "ساعات منتصف الليل",
    "—",
    "ضع 00:00 → 23:59",
    "—",
    "يُقبل",
    "", "P3-Low", "Not Started", "")

add("EM-065", "Positive", "Break", "إضافة استراحة",
    "يوم مفعّل",
    "اضغط (إضافة استراحة)",
    "—",
    "UI يضيف صف break + يظهر زر حذف",
    "نجح — UI يعمل.", "P2-Medium", "Pass",
    "Breaks تُحفظ كـ stub 200 بعد الإصلاح.")

add("EM-066", "Positive", "Break", "حذف استراحة",
    "استراحة موجودة",
    "اضغط (حذف) بجانبها",
    "—",
    "تختفي من UI",
    "", "P3-Low", "Not Started", "")

add("EM-067", "Negative", "Break", "استراحة خارج ساعات العمل",
    "يوم 09-17",
    "أضف استراحة 18:00-19:00",
    "—",
    "رسالة خطأ",
    "", "P2-Medium", "Not Started",
    "لا validation حالياً.")

add("EM-068", "Negative", "Break", "استراحتان متداخلتان",
    "—",
    "أضف 12-13 ثم 12:30-13:30",
    "—",
    "رسالة خطأ",
    "", "P2-Medium", "Not Started", "")

add("EM-069", "Positive", "Break", "عدة استراحات في يوم",
    "يوم 08-20",
    "أضف 3 استراحات (12-13, 16-17, 19-20)",
    "—",
    "الكل يُقبل",
    "", "P3-Low", "Not Started", "")

add("EM-070", "Positive", "Schedule", "Vacation switch (Tab 2 top)",
    "—",
    "فعّل switch (إجازة الممارس)",
    "—",
    "تظهر حقول تواريخ",
    "", "P2-Medium", "Not Started", "")

# ═══════════════════════════════════════════════════════════════════════════
# CATEGORY 6: CREATE — Services  (EM-071 → EM-085)
# ═══════════════════════════════════════════════════════════════════════════
add("EM-071", "Positive", "Service", "ربط خدمة واحدة",
    "خدمة موجودة",
    "tab 3 → اختر خدمة → 300/45 → أضف",
    "in_person: 300/45",
    "Draft يُضاف، POST يُرسل عند الحفظ",
    "نجح (POST 201).", "P1-High", "Pass", "")

add("EM-072", "Positive", "Service", "سعر مختلف لكل نوع حجز",
    "—",
    "حضوري 300/45، عن بُعد 200/30",
    "types=[{in_person:300/45},{online:200/30}]",
    "types[] في payload",
    "POST 201 لكن البيانات تُفقد (schema bug).",
    "P0-Critical", "Fail",
    "BUG EM-100: schema يحتاج EmployeeServiceType table.")

add("EM-073", "Positive", "Service", "buffer بين المواعيد",
    "—",
    "أضف bufferMinutes=15",
    "buffer=15",
    "يُحفظ",
    "نُرسل لكن يُفقد — نفس السبب.",
    "P1-High", "Fail", "BUG schema.")

add("EM-074", "Positive", "Service", "خدمات متعددة",
    "2 خدمات في DB",
    "أضف خدمتين بأسعار مختلفة",
    "—",
    "2 POST /services",
    "", "P2-Medium", "Not Started", "")

add("EM-075", "Negative", "Service", "نفس الخدمة مرتين",
    "خدمة مربوطة بالفعل",
    "حاول إضافتها مرة ثانية",
    "—",
    "409 Conflict",
    "", "P1-High", "Not Started",
    "تحقّق unique constraint.")

add("EM-076", "Negative", "Validation", "سعر سالب",
    "—",
    "أدخل price=-100",
    "—",
    "رسالة خطأ",
    "", "P2-Medium", "Not Started", "")

add("EM-077", "Negative", "Validation", "مدة = 0",
    "—",
    "أدخل duration=0",
    "—",
    "رسالة خطأ (min=1)",
    "", "P2-Medium", "Not Started", "")

add("EM-078", "Edge case", "Validation", "مدة > 8 ساعات",
    "—",
    "أدخل duration=600",
    "—",
    "مقبول أو warning",
    "", "P3-Low", "Not Started", "")

add("EM-079", "Positive", "Service", "نوع حجز واحد فقط",
    "—",
    "فعّل حضوري فقط، عطّل عن بُعد",
    "availableTypes=[in_person]",
    "types[] فيه عنصر واحد",
    "", "P2-Medium", "Not Started", "")

add("EM-080", "Positive", "Service", "Custom duration switch",
    "—",
    "فعّل (استخدام خيارات مدة مخصصة)",
    "useCustomOptions=true",
    "يعرض input لخيارات متعددة",
    "", "P2-Medium", "Not Started", "")

add("EM-081", "Positive", "Service", "isActive في الخدمة",
    "—",
    "عطّل الخدمة قبل الحفظ",
    "isActive=false",
    "payload.isActive=false",
    "", "P2-Medium", "Not Started", "")

add("EM-082", "Negative", "Dropdown", "خدمة بدون اسم",
    "خدمة nameAr=null في DB",
    "افتح dropdown",
    "—",
    "لا تظهر (skip) أو fallback للـ nameEn",
    "نجح بعد الإصلاح.", "P3-Low", "Pass",
    "Fixed: add-service-form.tsx")

add("EM-083", "Positive", "Service UX", "Cancel form الخدمة",
    "form مفتوح",
    "اضغط (إلغاء)",
    "—",
    "لا إضافة + إغلاق form",
    "نجح.", "P3-Low", "Pass", "")

add("EM-084", "Positive", "Service UX", "لا خدمات → empty state",
    "ممارس جديد",
    "tab 3",
    "—",
    "(لا توجد خدمات مرتبطة بعد)",
    "نجح.", "P3-Low", "Pass", "")

add("EM-085", "Edge case", "Service", "5+ خدمات",
    "5+ خدمات في DB",
    "أضف 5 خدمات",
    "—",
    "الكل يُحفظ، UI يعرضها",
    "", "P3-Low", "Not Started", "")

# ═══════════════════════════════════════════════════════════════════════════
# CATEGORY 7: DETAIL + EDIT  (EM-086 → EM-100)
# ═══════════════════════════════════════════════════════════════════════════
add("EM-086", "Positive", "Detail", "فتح صفحة التفاصيل",
    "ممارس موجود",
    "اضغط معاينة",
    "—",
    "GET /:id → 200، عرض جميع الحقول",
    "نجح.", "P1-High", "Pass", "")

add("EM-087", "Positive", "Detail", "عرض أوقات العمل",
    "availability محفوظة",
    "scroll لقسم (أوقات العمل)",
    "—",
    "عرض كل الأيام المفعّلة",
    "نجح.", "P2-Medium", "Pass", "")

add("EM-088", "Positive", "Detail", "عرض الخدمات المرتبطة",
    "خدمة مربوطة",
    "scroll لقسم (الخدمات المتاحة)",
    "—",
    "عرض اسم الخدمة + أسعار",
    "جزئي: الاسم يظهر، الأسعار `—` (schema).",
    "P1-High", "Fail", "BUG schema.")

add("EM-089", "Positive", "Detail", "زر تعديل من detail",
    "في detail",
    "اضغط (تعديل)",
    "—",
    "/employees/:id/edit",
    "نجح.", "P2-Medium", "Pass", "")

add("EM-090", "Negative", "Detail", "ممارس غير موجود",
    "id خطأ",
    "افتح /employees/fake-uuid",
    "—",
    "404 page أو رسالة خطأ",
    "", "P2-Medium", "Not Started", "")

add("EM-091", "Positive", "Edit", "prefill الحقول",
    "ممارس موجود",
    "افتح /edit",
    "—",
    "كل الحقول معبأة بالقيم الحالية",
    "نجح.", "P1-High", "Pass", "")

add("EM-092", "Positive", "Edit", "تعديل حقل واحد",
    "—",
    "غيّر اللقب فقط",
    "title=بروفيسور",
    "PATCH بـ partial payload → 200",
    "نجح.", "P1-High", "Pass", "")

add("EM-093", "Positive", "Edit", "تعديل كل الحقول",
    "—",
    "غيّر كل الحقول",
    "—",
    "PATCH full payload → 200",
    "", "P2-Medium", "Not Started", "")

add("EM-094", "Positive", "Edit", "Toast بعد الحفظ",
    "—",
    "احفظ",
    "—",
    "Toast نجاح + redirect",
    "نجح بعد الإصلاح.", "P1-High", "Pass",
    "Fixed: use-employee-form.ts — non-blocking side effects.")

add("EM-095", "Positive", "Edit", "Cancel بدون حفظ",
    "تعديلات غير محفوظة",
    "اضغط (إلغاء)",
    "—",
    "redirect بدون PATCH",
    "نجح.", "P2-Medium", "Pass", "")

add("EM-096", "Edge case", "Edit", "تعديل بعد اختفاء الممارس",
    "ممارس محذوف من DB",
    "احفظ تعديل",
    "—",
    "404 + رسالة واضحة",
    "", "P3-Low", "Not Started", "")

add("EM-097", "Negative", "Edit", "PATCH بـ email مكرر",
    "ممارس آخر بنفس الإيميل",
    "غيّر إيميل → نفس إيميل موجود",
    "—",
    "409",
    "", "P1-High", "Not Started",
    "UPDATE DTO لا يقبل email حالياً — يحتاج فحص.")

add("EM-098", "Positive", "Edit", "القائمة تحدّث بعد EDIT",
    "ممارس معدّل",
    "ارجع للقائمة",
    "—",
    "القيم الجديدة تظهر",
    "نجح — refetchType:all.", "P1-High", "Pass",
    "Fixed: use-employee-mutations.ts.")

add("EM-099", "Positive", "Edit", "Stats تحدّث بعد isActive flip",
    "ممارس نشط",
    "عطّله → احفظ",
    "isActive=false",
    "Stats: نشط -=1, غير متاح +=1",
    "", "P1-High", "Not Started", "")

add("EM-100", "Positive", "Edit", "تعديل schedule",
    "—",
    "غيّر ساعات يوم + احفظ",
    "—",
    "PATCH /availability → 200 + detail يعكسها",
    "", "P1-High", "Not Started",
    "spinbutton صعب الاختبار.")

# ═══════════════════════════════════════════════════════════════════════════
# CATEGORY 8: DELETE  (EM-101 → EM-108)
# ═══════════════════════════════════════════════════════════════════════════
add("EM-101", "Positive", "Delete", "حذف ناجح",
    "ممارس موجود",
    "أيقونة حذف → تأكيد",
    "—",
    "DELETE → 204 + اختفاء + stats تحدّث",
    "نجح.", "P0-Critical", "Pass",
    "Fixed: delete-employee-dialog.tsx + api.ts (handle 204).")

add("EM-102", "Positive", "Delete", "إلغاء التأكيد",
    "dialog مفتوح",
    "اضغط (إلغاء)",
    "—",
    "لا DELETE",
    "نجح.", "P2-Medium", "Pass", "")

add("EM-103", "Positive", "Delete", "Toast نجاح",
    "—",
    "احذف",
    "—",
    "Toast (تم حذف الممارس)",
    "نجح.", "P2-Medium", "Pass", "")

add("EM-104", "Positive", "Delete", "Stats تحدّث فوراً",
    "Stats قبل: total=5",
    "احذف واحد",
    "—",
    "total=4 فوراً",
    "نجح — invalidate stats key.", "P1-High", "Pass", "")

add("EM-105", "Negative", "Delete", "ممارس له حجوزات مستقبلية",
    "ممارس + booking قائم",
    "احذفه",
    "—",
    "رفض أو warning (cascade behavior)",
    "", "P0-Critical", "Not Started",
    "تحقّق onDelete في Booking schema.")

add("EM-106", "Negative", "Delete", "ممارس له تقييمات",
    "ممارس + ratings",
    "احذفه",
    "—",
    "cascade مناسب",
    "", "P1-High", "Not Started", "")

add("EM-107", "Edge case", "Delete", "حذف متتالي سريع",
    "3 ممارسين",
    "احذف 3 خلال ثانية",
    "—",
    "الكل يُحذف + Stats صحيحة",
    "", "P3-Low", "Not Started", "")

add("EM-108", "Negative", "Delete", "ممارس من tenant آخر",
    "2 tenants",
    "حاول حذف ممارس من tenant B عبر curl",
    "—",
    "404 (tenant isolation)",
    "", "P0-Critical", "Not Started",
    "تحقّق WHERE tenantId في delete handler.")

# ═══════════════════════════════════════════════════════════════════════════
# CATEGORY 9: VACATIONS  (EM-109 → EM-114)
# ═══════════════════════════════════════════════════════════════════════════
add("EM-109", "Positive", "Vacation", "إضافة إجازة",
    "في edit",
    "فعّل switch → حدّد تواريخ → احفظ",
    "2026-05-01 → 2026-05-10",
    "POST /vacations → 201",
    "نجح بعد الإصلاح.", "P1-High", "Pass",
    "Added: endpoint alias.")

add("EM-110", "Positive", "Vacation", "عرض الإجازات",
    "إجازة محفوظة",
    "detail",
    "—",
    "GET /vacations → 200 + عرض",
    "نجح.", "P2-Medium", "Pass", "")

add("EM-111", "Positive", "Vacation", "حذف إجازة",
    "—",
    "اضغط حذف بجانب الإجازة",
    "—",
    "DELETE → 204",
    "نجح.", "P2-Medium", "Pass", "")

add("EM-112", "Negative", "Vacation", "endDate < startDate",
    "—",
    "ضع بداية 2026-05-10، نهاية 2026-05-01",
    "—",
    "رفض",
    "", "P1-High", "Not Started", "")

add("EM-113", "Negative", "Vacation", "إجازات متداخلة",
    "إجازة 1-10",
    "أضف إجازة 5-15",
    "—",
    "warning أو رفض",
    "", "P2-Medium", "Not Started", "")

add("EM-114", "Edge case", "Vacation", "إجازة في الماضي",
    "—",
    "بداية 2020-01-01",
    "—",
    "رفض أو warning",
    "", "P3-Low", "Not Started", "")

# ═══════════════════════════════════════════════════════════════════════════
# CATEGORY 10: AUTH + AUTHORIZATION  (EM-115 → EM-120)
# ═══════════════════════════════════════════════════════════════════════════
add("EM-115", "Negative", "Auth", "بدون JWT",
    "logged out",
    "curl GET /employees",
    "—",
    "401",
    "", "P0-Critical", "Not Started",
    "تحقّق JwtGuard.")

add("EM-116", "Negative", "Auth", "JWT منتهي",
    "token منتهي",
    "أي طلب",
    "—",
    "401 → refresh → retry",
    "نجح (أمس رأيت POST /auth/refresh يعمل).",
    "P1-High", "Pass", "")

add("EM-117", "Negative", "Authz", "RECEPTIONIST يحذف",
    "role=RECEPTIONIST",
    "حاول DELETE",
    "—",
    "403",
    "", "P0-Critical", "Not Started",
    "تحقّق CaslGuard.")

add("EM-118", "Negative", "Authz", "Tenant isolation",
    "2 tenants",
    "list employees من tenant A بـ token من B",
    "—",
    "قائمة B فقط",
    "", "P0-Critical", "Not Started", "")

add("EM-119", "Negative", "Auth", "token مزوّر",
    "—",
    "عدّل token",
    "—",
    "401",
    "", "P0-Critical", "Not Started", "")

add("EM-120", "Positive", "Auth", "Topbar يعرض اسم المستخدم",
    "admin مسجّل",
    "—",
    "—",
    "(A — Admin — مدير العيادة) وليس ??",
    "نجح.", "P1-High", "Pass",
    "Fixed: get-current-user + header fallback.")

# ═══════════════════════════════════════════════════════════════════════════
# BUILD SHEET
# ═══════════════════════════════════════════════════════════════════════════
wb = openpyxl.load_workbook(WB_PATH)
src = wb["E2E Test Cases"]
hdr = [src.cell(row=1, column=c) for c in range(1, src.max_column + 1)]

if SHEET in wb.sheetnames:
    del wb[SHEET]
ws = wb.create_sheet(SHEET)

for i, h in enumerate(HEADERS, 1):
    c = ws.cell(row=1, column=i, value=h)
    s = hdr[i-1]
    c.font = copy(s.font); c.fill = copy(s.fill)
    c.alignment = copy(s.alignment); c.border = copy(s.border)

widths = [10, 22, 14, 22, 40, 30, 35, 25, 40, 40, 10, 13, 12, 12, 12, 35]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

ws.row_dimensions[1].height = 30
ws.freeze_panes = "A2"

thin = Side(style="thin", color="CCCCCC")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
fill_pass = PatternFill("solid", fgColor="D4EDDA")
fill_fail = PatternFill("solid", fgColor="F8D7DA")
fill_block = PatternFill("solid", fgColor="FFF3CD")

def is_ar(s):
    return any("\u0600" <= ch <= "\u06ff" for ch in (s or ""))

for row_idx, row in enumerate(T, start=2):
    tid, ttype, scenario, title, pre, steps, data, expected, actual, priority, status, notes = row
    cells = [tid, MODULE, ttype, scenario, title, pre, steps, data, expected,
             actual, priority, status, ENV, EXEC if status != "Not Started" else "",
             TODAY if status != "Not Started" else "", notes]
    for col_idx, val in enumerate(cells, start=1):
        c = ws.cell(row=row_idx, column=col_idx, value=val)
        text = val if isinstance(val, str) else ""
        c.alignment = Alignment(
            wrap_text=True, vertical="top",
            horizontal="right" if is_ar(text) else "left")
        c.border = border
        if col_idx == 12:
            if val == "Pass": c.fill = fill_pass
            elif val == "Fail": c.fill = fill_fail
            elif val == "Blocked": c.fill = fill_block
    ws.row_dimensions[row_idx].height = 72

ws.sheet_view.rightToLeft = True

wb.save(WB_PATH)

# Stats
counts = {}
for row in T:
    counts[row[10]] = counts.get(row[10], 0) + 1
print(f"Total: {len(T)} cases")
for k, v in sorted(counts.items()): print(f"  {k}: {v}")

status_counts = {}
for row in T:
    status_counts[row[10]] = status_counts.get(row[10], 0)
    s = row[10]
for row in T:
    s = row[10]
# Actually count status
sc = {}
for row in T:
    sc[row[10]] = sc.get(row[10], 0) + 1
print("\nBy Priority:", sc)

by_status = {}
for row in T:
    by_status[row[10]] = by_status.get(row[10], 0) + 1

by_st = {}
for row in T:
    st = row[10]  # this is priority
by_st = {}
for row in T:
    st = row[10]
# redo
from collections import Counter
print("By Status:", Counter(r[10] for r in T))
print("By Status (real):", Counter(r[10] for r in T))
# Status is at index 10? Let me recheck. Row tuple:
# 0 tid, 1 ttype, 2 scenario, 3 title, 4 pre, 5 steps, 6 data, 7 expected, 8 actual, 9 priority, 10 status, 11 notes
print("\nStatus distribution:", Counter(r[10] for r in T))
print("Priority distribution:", Counter(r[9] for r in T))
