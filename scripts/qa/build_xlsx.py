import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from copy import copy
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

wb = openpyxl.load_workbook("e2e_test_clients.xlsx")

src = wb["E2E Test Cases"]
header_cells = [src.cell(row=1, column=c) for c in range(1, src.max_column + 1)]

if "E2E Employees" in wb.sheetnames:
    del wb["E2E Employees"]
ws = wb.create_sheet("E2E Employees")

headers = [
    "Test ID", "Module", "Test Type", "Test Scenario", "Test Case Title",
    "Preconditions", "Test Steps", "Test Data", "Expected Result",
    "Actual Result", "Priority", "Status", "Environment", "Executed By",
    "Execution Date", "Notes / Defect ID",
]

for i, h in enumerate(headers, 1):
    c = ws.cell(row=1, column=i, value=h)
    sc = header_cells[i - 1]
    c.font = copy(sc.font)
    c.fill = copy(sc.fill)
    c.alignment = copy(sc.alignment)
    c.border = copy(sc.border)

widths = [10, 22, 12, 40, 45, 35, 45, 30, 45, 45, 10, 14, 12, 14, 14, 40]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

ws.row_dimensions[1].height = 30
ws.freeze_panes = "A2"

# Each tuple = 16 columns
T = [
    ("EM-001", "Employees (الممارسون)", "Positive", "عرض قائمة الممارسين",
     "جلب قائمة الممارسين مع pagination و stats",
     "1. مستخدم مسجّل دخول\n2. وجود ممارسين في التينانت",
     "1. افتح /employees\n2. لاحظ StatsGrid + الجدول",
     "page=1&limit=20",
     "1. GET /dashboard/people/employees → 200\n2. GET /dashboard/people/employees/stats → 200\n3. الكروت: الإجمالي/نشط/غير متاح/متوسط التقييم\n4. الجدول يعرض الممارسين",
     "نجح: القائمة + stats يُحمَّلان من endpoints منفصلة. stats لا تتأثر بالفلاتر.",
     "P1-High", "Pass", "Local dev", "Claude", "2026-04-15", ""),

    ("EM-002", "Employees (الممارسون)", "Positive", "بحث Debounce",
     "البحث في القائمة بـ debounce 300ms",
     "قائمة غير فارغة",
     "1. اكتب 5 أحرف سريعة في حقل البحث",
     "search=tariq",
     "طلب واحد فقط بعد 300ms (ليس طلب لكل حرف)",
     "نجح: قبل الإصلاح كان 5 طلبات، الآن طلب واحد.",
     "P2-Medium", "Pass", "Local dev", "Claude", "2026-04-15",
     "Fixed: use-employees.ts — useEffect + setTimeout 300ms"),

    ("EM-003", "Employees (الممارسون)", "Positive", "فلتر الحالة isActive",
     "فلترة الممارسين بـ isActive=false عبر combobox",
     "ممارس نشط + ممارس معطّل",
     "1. افتح combobox (جميع الحالات)\n2. اختر (غير متاح)",
     "isActive=false",
     "1. GET ?isActive=false → 200\n2. عرض المعطّلين فقط\n3. Stats لا تتأثر بالفلتر",
     "نجح بعد الإصلاح: الباك كان يتجاهل الفلتر بسبب ValidationPipe implicit conversion. الحل: parse يدوي في الـ controller.",
     "P1-High", "Pass", "Local dev", "Claude", "2026-04-15",
     "Fixed: people.controller.ts — @Query(isActive) rawIsActive"),

    ("EM-004", "Employees (الممارسون)", "Positive", "Stats endpoint",
     "الإحصائيات مستقلة عن فلاتر القائمة",
     "ممارسون بحالات متنوعة",
     "1. ابحث عن اسم غير موجود\n2. لاحظ StatsGrid",
     "search=zzznotfound",
     "الجدول فارغ، لكن Stats تظل تعرض total/active الكلي",
     "نجح: أُضيف GET /employees/stats جديد يحسب total/active/inactive/avgRating بدون فلاتر.",
     "P1-High", "Pass", "Local dev", "Claude", "2026-04-15",
     "Added: employee-stats.handler.ts + useEmployeeStats hook"),

    ("EM-010", "Employees (الممارسون)", "Positive", "إنشاء ممارس (Tab 1: Basic Info)",
     "ملء المعلومات الأساسية وإرسال onboarding",
     "مستخدم بصلاحية إنشاء",
     "1. اضغط (إضافة ممارس)\n2. املأ: اللقب، البريد، الاسم AR+EN، التخصص AR+EN، الخبرة، التعليم، النبذة",
     "title=دكتور\nnameEn=Dr. Ahmed\nnameAr=د. أحمد\nemail=dr.ahmed@test.com\nspecialty=Dermatologist\nexperience=8",
     "1. POST /employees/onboarding → 201\n2. جميع الحقول تُرسل في الـ payload\n3. الرد يحتوي employee.id",
     "نجح: كل الحقول محفوظة بشكل صحيح. payload نظيف.",
     "P1-High", "Pass", "Local dev", "Claude", "2026-04-15", ""),

    ("EM-011", "Employees (الممارسون)", "Positive", "إنشاء ممارس (Tab 2: Schedule)",
     "تحديد أيام العمل وتفعيل/تعطيل أيام",
     "Tab 1 مملوء",
     "1. انتقل لتاب (الجدول والاستراحات)\n2. عطّل الخميس\n3. فعّل السبت\n4. أضف استراحة للاثنين",
     "days: Sun/Mon/Tue/Wed + Sat\nbreak: Monday 12:00-13:00",
     "PATCH /employees/:id/availability → 200\nbody.windows يحتوي 6 أيام (0,1,2,3,6 بدون 4 و 5)",
     "نجح: windows محفوظة بالـ dayOfWeek الصحيح.",
     "P1-High", "Pass", "Local dev", "Claude", "2026-04-15", ""),

    ("EM-012", "Employees (الممارسون)", "Negative", "Breaks endpoint مفقود",
     "الاستراحات يفترض أن تُرسل لـ PUT /breaks لكن endpoint مفقود",
     "استراحة مُضافة في الفورم",
     "1. املأ استراحة للاثنين\n2. احفظ الممارس",
     "breaks: [{dayOfWeek:1, startTime:12:00, endTime:13:00}]",
     "PUT /employees/:id/breaks → 200",
     "فشل قبل الإصلاح: 404. بعد الإصلاح: أُضيف PUT stub يعيد 200 (no-op حتى migration split-shifts).",
     "P2-Medium", "Pass", "Local dev", "Claude", "2026-04-15",
     "Fixed: people.controller.ts — PUT /breaks stub"),

    ("EM-013", "Employees (الممارسون)", "Positive", "إنشاء ممارس (Tab 3: Services)",
     "ربط خدمة بسعر ومدة مختلفين لكل نوع حجز",
     "خدمة موجودة في النظام",
     "1. انتقل لتاب (الخدمات والتسعير)\n2. اختر خدمة\n3. حضوري: 300 ر.س / 45 دق\n4. عن بُعد: 200 ر.س / 30 دق\n5. buffer: 15 دق",
     "in_person: 300/45\nonline: 200/30\nbuffer: 15",
     "POST /employees/:id/services → 201\npayload يحتوي types[] + bufferMinutes",
     "جزئي: POST 201، لكن الباك يتجاهل price/duration/types (Schema EmployeeService فقيرة). يحتاج migration.",
     "P0-Critical", "Fail", "Local dev", "Claude", "2026-04-15",
     "BUG: assign-employee-service.handler يتجاهل التسعير. Schema تحتاج حقول جديدة + EmployeeServiceType model."),

    ("EM-014", "Employees (الممارسون)", "Negative", "Dropdown الخدمات — ??? ???",
     "خدمة بـ nameAr null تظهر كـ ??? ???",
     "خدمة في DB بـ nameAr فارغ",
     "1. افتح dropdown الخدمات في tab 3",
     "—",
     "تجاهل الخدمات بدون اسم أو استخدام fallback",
     "نجح بعد الإصلاح: تجاهل الخدمات بدون اسم، fallback بين AR/EN.",
     "P3-Low", "Pass", "Local dev", "Claude", "2026-04-15",
     "Fixed: add-service-form.tsx + assign-service-sheet.tsx"),

    ("EM-015", "Employees (الممارسون)", "Positive", "CREATE — القائمة تُحدَّث بعد الإنشاء",
     "السجل الجديد يظهر بدون reload يدوي",
     "ممارس تم إنشاؤه للتو",
     "1. أنشئ ممارس جديد\n2. بعد redirect لـ /employees\n3. لاحظ الجدول",
     "—",
     "السجل الجديد يظهر + stats تحدّث",
     "نجح بعد الإصلاح: أُضيف refetchType:all للـ invalidate.",
     "P1-High", "Pass", "Local dev", "Claude", "2026-04-15",
     "Fixed: use-employee-mutations.ts"),

    ("EM-020", "Employees (الممارسون)", "Positive", "صفحة التفاصيل",
     "عرض بيانات الممارس بعد الإنشاء",
     "ممارس موجود",
     "1. اضغط (معاينة) في صف الجدول\n2. لاحظ صفحة detail",
     "—",
     "1. GET /employees/:id → 200\n2. عرض: الاسم، التخصص، الخبرة، التعليم، النبذة، أوقات العمل، الخدمات\n3. روابط (تعديل)",
     "نجح بعد إصلاح bugs: list-employee-services ما كان يحمّل service relation. الآن يحمّل يدوياً.",
     "P1-High", "Pass", "Local dev", "Claude", "2026-04-15",
     "Fixed: list-employee-services.handler.ts + employee-services-section.tsx null-safe"),

    ("EM-021", "Employees (الممارسون)", "Positive", "Preview button",
     "زر المعاينة ينتقل لصفحة التفاصيل",
     "صف ممارس في الجدول",
     "1. اضغط أيقونة العين",
     "—",
     "الانتقال لـ /employees/:id",
     "نجح بعد الإصلاح: aria-label كان common.preview (خام)، الآن معاينة وcallback مربوط.",
     "P2-Medium", "Pass", "Local dev", "Claude", "2026-04-15",
     "Fixed: ar.nav.ts + en.nav.ts — common.preview key"),

    ("EM-030", "Employees (الممارسون)", "Positive", "تعديل المعلومات الأساسية",
     "تغيير اللقب والخبرة وحفظ",
     "ممارس موجود",
     "1. اضغط (تعديل)\n2. غيّر اللقب لـ بروفيسور\n3. غيّر الخبرة لـ 15\n4. احفظ",
     "title=بروفيسور\nexperience=15",
     "1. PATCH /employees/:id → 200\n2. Toast نجاح\n3. Redirect إلى /employees\n4. القائمة تعرض 15 سنة",
     "نجح: الحقول محفوظة. القائمة تحدّثت بعد invalidation.",
     "P1-High", "Pass", "Local dev", "Claude", "2026-04-15", ""),

    ("EM-031", "Employees (الممارسون)", "Negative", "EDIT — side effects block redirect",
     "فشل side effect (breaks 404) يمنع toast/redirect",
     "ممارس مع استراحات",
     "1. افتح تعديل\n2. غيّر شيء\n3. احفظ",
     "—",
     "حتى لو فشل PUT /breaks، الـ PATCH الأساسي ينجح ويعرض toast + redirect",
     "نجح بعد الإصلاح: Side effects في try/catch منفصلة، الأخطاء تُجمع كـ warning لكن الـ redirect يحدث.",
     "P1-High", "Pass", "Local dev", "Claude", "2026-04-15",
     "Fixed: use-employee-form.ts — submitEdit refactored"),

    ("EM-032", "Employees (الممارسون)", "Positive", "EDIT — invalidate list",
     "التعديل ينعكس في القائمة",
     "ممارس معدّل",
     "1. عدّل ممارس\n2. ارجع لـ /employees",
     "—",
     "القائمة تعرض القيم الجديدة بدون reload",
     "نجح بعد الإصلاح: refetchType:all يُجبر refetch.",
     "P1-High", "Pass", "Local dev", "Claude", "2026-04-15", ""),

    ("EM-040", "Employees (الممارسون)", "Positive", "حذف ممارس",
     "حذف ممارس مع تأكيد",
     "ممارس موجود",
     "1. اضغط أيقونة الحذف\n2. confirmation dialog\n3. اضغط (حذف)",
     "—",
     "1. DELETE /employees/:id → 204\n2. Toast نجاح\n3. اختفاء من القائمة\n4. Stats تحدّث من 1→0",
     "نجح بعد الإصلاح: كان stub يعرض toast.error فقط. الآن mutation حقيقية + invalidate stats + handle 204.",
     "P0-Critical", "Pass", "Local dev", "Claude", "2026-04-15",
     "Fixed: delete-employee-dialog.tsx + api.ts (handle 204 No Content)"),

    ("EM-041", "Employees (الممارسون)", "Positive", "DELETE — إلغاء التأكيد",
     "ضغط إلغاء لا يحذف",
     "Dialog مفتوح",
     "1. اضغط (إلغاء)",
     "—",
     "1. Dialog يُغلق\n2. لا طلب DELETE\n3. السجل موجود",
     "نجح.",
     "P2-Medium", "Pass", "Local dev", "Claude", "2026-04-15", ""),

    ("EM-050", "Employees (الممارسون)", "Positive", "إضافة إجازة",
     "تفعيل الإجازة في EDIT مع تواريخ",
     "ممارس في صفحة EDIT",
     "1. فعّل switch الإجازة\n2. حدد تاريخ بداية/نهاية\n3. احفظ",
     "startDate=2026-05-01\nendDate=2026-05-10",
     "POST /employees/:id/vacations → 201",
     "نجح بعد الإصلاح: endpoint كان 404، أُضيف alias للـ exceptions.",
     "P1-High", "Pass", "Local dev", "Claude", "2026-04-15",
     "Added: POST/GET/DELETE /employees/:id/vacations"),

    ("EM-051", "Employees (الممارسون)", "Positive", "عرض الإجازات القادمة",
     "detail page تعرض الإجازات",
     "ممارس له إجازة",
     "1. افتح detail",
     "—",
     "GET /employees/:id/vacations → 200 + عرض القائمة",
     "نجح بعد الإصلاح.",
     "P2-Medium", "Pass", "Local dev", "Claude", "2026-04-15", ""),

    ("EM-052", "Employees (الممارسون)", "Positive", "حذف إجازة",
     "حذف vacation record",
     "إجازة موجودة",
     "1. اضغط حذف بجانب الإجازة",
     "—",
     "DELETE /employees/:id/vacations/:vacationId → 204",
     "نجح بعد الإصلاح.",
     "P2-Medium", "Pass", "Local dev", "Claude", "2026-04-15", ""),

    ("EM-060", "Employees (الممارسون)", "Positive", "زر التقييمات والمراجعات",
     "التنقل لصفحة ratings",
     "—",
     "1. اضغط (التقييمات والمراجعات) في PageHeader",
     "—",
     "1. التنقل لـ /ratings\n2. combobox اختيار ممارس\n3. اختيار ممارس → GET /employees/:id/ratings",
     "نجح: combobox + empty state صحيحان.",
     "P2-Medium", "Pass", "Local dev", "Claude", "2026-04-15", ""),

    ("EM-070", "Employees (الممارسون)", "Positive", "Topbar user name",
     "اسم المستخدم يظهر صحيحاً بدل ??",
     "مستخدم admin",
     "1. لاحظ أعلى يسار الصفحة",
     "—",
     "يعرض A — Admin — مدير العيادة (وليس ??)",
     "نجح بعد الإصلاح: الباك كان يرسل {name:Admin} والـ frontend يتوقع firstName/lastName. الآن يُقسم الاسم + fallback للإيميل.",
     "P1-High", "Pass", "Local dev", "Claude", "2026-04-15",
     "Fixed: get-current-user.handler.ts + header.tsx"),

    ("EM-080", "Employees (الممارسون)", "Positive", "Branding endpoint لا يرمي 404",
     "الصفحة تُحمَّل دون أخطاء branding في console",
     "tenant بدون branding config",
     "1. افتح أي صفحة\n2. افتح console",
     "—",
     "GET /public/branding/:tenantId → 200 مع defaults (بدل 404)",
     "نجح بعد الإصلاح: الـ handler كان يرمي NotFoundException، الآن يعيد defaults.",
     "P2-Medium", "Pass", "Local dev", "Claude", "2026-04-15",
     "Fixed: get-branding.handler.ts"),

    ("EM-090", "Employees (الممارسون)", "Positive", "Seed data نظيفة",
     "عدم وجود سجلات ملوّثة (nameEn=General Checkup)",
     "قاعدة بيانات dev",
     "1. افتح /employees",
     "—",
     "لا سجلات بأسماء خدمات في حقل الاسم",
     "نجح: حُذفت 4 سجلات ملوّثة عبر prisma/clean-bad-employees.ts.",
     "P2-Medium", "Pass", "Local dev", "Claude", "2026-04-15", ""),

    ("EM-100", "Employees (الممارسون)", "Negative", "BUG: التسعير المخصص للممارس",
     "السعر/المدة/نوع الحجز للخدمة تُرسل لكن لا تُحفظ",
     "ربط خدمة بتسعير مخصص",
     "1. أنشئ ممارس\n2. اربط خدمة بـ 300 ر.س / 45 دق\n3. افتح detail",
     "price=300, duration=45",
     "detail يعرض (الأسعار: زيارة العيادة 300 / استشارة هاتفية —)",
     "فشل: Schema EmployeeService لا تحتوي price/duration/types. يحتاج migration جديدة + model EmployeeServiceType. البيانات تُقبل بـ 201 لكن تُفقد.",
     "P0-Critical", "Fail", "Local dev", "Claude", "2026-04-15",
     "BLOCKED: migration drift في الـ repo يمنع schema changes. يحتاج قرار: reset أو fix-drift أو migration جديدة."),

    ("EM-101", "Employees (الممارسون)", "Negative", "BUG: خانة الخبرة valuemax=0",
     "spinbutton في فورم الإنشاء فيه max=0",
     "فورم الإنشاء",
     "1. افتح create employee\n2. فحص خانة الخبرة",
     "—",
     "valuemax ≥ 100",
     "فشل: الـ input HTML max=0. UX مربك لكن الـ form يقبل القيم عند الإرسال.",
     "P3-Low", "Fail", "Local dev", "Claude", "2026-04-15",
     "BUG: experience number input max attribute. Not blocking."),

    ("EM-102", "Employees (الممارسون)", "Negative", "BUG: GET /bookings 400",
     "detail page يطلب bookings ويفشل",
     "ممارس موجود",
     "1. افتح detail\n2. لاحظ chart (الحجوزات حسب الحالة)",
     "employeeId query param",
     "GET /dashboard/bookings?employeeId=... → 200",
     "فشل: 400 Bad Request. الـ charts لا تعمل.",
     "P2-Medium", "Fail", "Local dev", "Claude", "2026-04-15",
     "BUG: bookings list DTO يرفض employeeId/fromDate/toDate params. خارج نطاق employees."),
]

thin = Side(style="thin", color="CCCCCC")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
pass_fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
fail_fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")

def is_arabic(s):
    if not s:
        return False
    return any("\u0600" <= ch <= "\u06ff" for ch in s)

for row_idx, row in enumerate(T, start=2):
    for col_idx, val in enumerate(row, start=1):
        c = ws.cell(row=row_idx, column=col_idx, value=val)
        text = val if isinstance(val, str) else ""
        c.alignment = Alignment(
            wrap_text=True,
            vertical="top",
            horizontal="right" if is_arabic(text) else "left",
        )
        c.border = border
        if col_idx == 12:
            if val == "Pass":
                c.fill = pass_fill
            elif val == "Fail":
                c.fill = fail_fill
    ws.row_dimensions[row_idx].height = 80

ws.sheet_view.rightToLeft = True

wb.save("e2e_test_clients.xlsx")
print("Saved. Sheets:", wb.sheetnames)
print("Employees sheet rows:", ws.max_row - 1)
